import os
import openpyxl as op

# File paths
erp_file_path = "C:\\Users\\Usefr\\Desktop\\TransferPrice\\erp.xlsx"
not_found_path = "C:\\Users\\Usefr\\Desktop\\TransferPrice\\Results\\NotFound.txt"
site_file_path = "C:\\Users\\Usefr\\Desktop\\TransferPrice\\site.txt"
final_xlsx_path = "C:\\Users\\Usefr\\Desktop\\TransferPrice\\Results\\FinalFile.xlsx"
main_dict_path = "C:\\Users\\Usefr\\Desktop\\TransferPrice\\Results"

# Load ERP workbook and sheet
workBook1 = op.load_workbook(erp_file_path)
workSheet1 = workBook1['Planilha1']

# Create a new workbook and add a worksheet
final_workbook = op.Workbook()
final_worksheet = final_workbook.active

# Create the main directory if it doesn't exist
os.makedirs(main_dict_path, exist_ok=True)

# Initialize the dictionary to store barcode and price data
codePrice = {}

# Read data from ERP workbook into codePrice dictionary
for row in workSheet1.iter_rows(min_row=1, max_row=workSheet1.max_row, min_col=1, max_col=2, values_only=True):
    bar_code = str(row[0])
    price = row[1]
    codePrice[bar_code] = price

# Close ERP workbook
workBook1.close()

# Read barcodes from the site file
with open(site_file_path, "r") as file:
    barCodes = [value.strip() for value in file.readlines()]

# Check each barcode
with open(not_found_path, "w+") as file:
    c = 0
    for value in barCodes:
        if value in codePrice:
            # If the barcode is found, append it to the final worksheet
            final_worksheet.append([value, codePrice[value]])
        else:
            # If the barcode is not found, write it to the NotFound.txt file
            final_worksheet.append([value, "Price Not Found"])
            file.write(f"O produto {value} não foi encontrado no ERP.")
            file.write("\n")
            c += 1
    if c == 0:
        file.write(f"All the products were found.")
    else:
        file.write(f"There's {c} products that weren't found.")

# Save the final workbook
final_workbook.save(final_xlsx_path)